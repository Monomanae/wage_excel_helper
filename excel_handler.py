"""Excel reading and writing for Wage Form Helper."""

import openpyxl
from openpyxl.utils import column_index_from_string
import config


def read_wage_form(file_path):
    """
    Read worker records from the wage input Excel file.
    Returns a list of dicts: {name, id, bank_account, bank_name, amount}
    Raises ValueError with a clear message if columns are missing.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    sheet_name = config.WAGE_FORM["sheet_name"]
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"工资表中未找到工作表 '{sheet_name}'\n"
            f"文件中的工作表: {wb.sheetnames}\n\n"
            f"Sheet '{sheet_name}' not found in the wage form.\n"
            f"Available sheets: {wb.sheetnames}\n\n"
            f"请更新 config.py 中的 WAGE_FORM['sheet_name']"
        )

    ws = wb[sheet_name]
    hdr_row = config.WAGE_FORM["header_row"]

    # Build a map: header_text → column_index
    header_map = {}
    for cell in ws[hdr_row]:
        if cell.value is not None:
            header_map[str(cell.value).strip()] = cell.column

    def resolve_col(config_key):
        col_label = config.WAGE_FORM[config_key]
        if col_label not in header_map:
            available = list(header_map.keys())
            raise ValueError(
                f"列 '{col_label}' 未找到 / Column '{col_label}' not found.\n"
                f"可用列 / Available columns: {available}\n\n"
                f"请更新 config.py 中的 '{config_key}'"
            )
        return header_map[col_label]

    col_name    = resolve_col("col_name")
    col_id      = resolve_col("col_id")
    col_bank_no = resolve_col("col_bank_account")
    col_bank_nm = resolve_col("col_bank_name")
    col_amount  = resolve_col("col_amount")

    workers = []
    for row_idx in range(hdr_row + 1, ws.max_row + 1):
        name_val = ws.cell(row=row_idx, column=col_name).value
        if name_val is None or str(name_val).strip() == "":
            continue  # skip blank rows

        workers.append({
            "name":         str(name_val).strip(),
            "id":           _str(ws.cell(row=row_idx, column=col_id).value),
            "bank_account": _str(ws.cell(row=row_idx, column=col_bank_no).value),
            "bank_name":    _str(ws.cell(row=row_idx, column=col_bank_nm).value),
            "amount":       _to_float(ws.cell(row=row_idx, column=col_amount).value),
        })

    return workers


def fill_bank_form(template_path, output_path, workers):
    """
    Copy bank template and write worker data into it.
    Saves the result to output_path; does NOT modify the template.
    """
    wb = openpyxl.load_workbook(template_path)

    sheet_name = config.BANK_FORM["sheet_name"]
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"银行模板中未找到工作表 '{sheet_name}'\n"
            f"文件中的工作表: {wb.sheetnames}\n\n"
            f"Sheet '{sheet_name}' not found in the bank template.\n"
            f"请更新 config.py 中的 BANK_FORM['sheet_name']"
        )

    ws = wb[sheet_name]
    start = config.BANK_FORM["data_start_row"]

    def col_idx(val):
        return val if isinstance(val, int) else column_index_from_string(str(val))

    c_name = col_idx(config.BANK_FORM["col_name"])
    c_id   = col_idx(config.BANK_FORM["col_id"])
    c_bno  = col_idx(config.BANK_FORM["col_bank_account"])
    c_bnm  = col_idx(config.BANK_FORM["col_bank_name"])
    c_amt  = col_idx(config.BANK_FORM["col_amount"])

    for i, w in enumerate(workers):
        row = start + i
        ws.cell(row=row, column=c_name).value = w.get("name", "")
        ws.cell(row=row, column=c_id).value   = w.get("id", "")
        ws.cell(row=row, column=c_bno).value  = w.get("bank_account", "")
        ws.cell(row=row, column=c_bnm).value  = w.get("bank_name", "")
        ws.cell(row=row, column=c_amt).value  = w.get("amount", 0)

    wb.save(output_path)


def read_bank_form(file_path):
    """
    Read back the data rows from an already-filled bank form.
    Used to show the user exactly what ended up in the bank form.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[config.BANK_FORM["sheet_name"]]
    start = config.BANK_FORM["data_start_row"]

    def col_idx(val):
        return val if isinstance(val, int) else column_index_from_string(str(val))

    c_name = col_idx(config.BANK_FORM["col_name"])
    c_id   = col_idx(config.BANK_FORM["col_id"])
    c_bno  = col_idx(config.BANK_FORM["col_bank_account"])
    c_bnm  = col_idx(config.BANK_FORM["col_bank_name"])
    c_amt  = col_idx(config.BANK_FORM["col_amount"])

    workers = []
    for row_idx in range(start, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=c_name).value
        if name is None or str(name).strip() == "":
            continue
        workers.append({
            "name":         _str(name),
            "id":           _str(ws.cell(row=row_idx, column=c_id).value),
            "bank_account": _str(ws.cell(row=row_idx, column=c_bno).value),
            "bank_name":    _str(ws.cell(row=row_idx, column=c_bnm).value),
            "amount":       _to_float(ws.cell(row=row_idx, column=c_amt).value),
        })
    return workers


# ── Helpers ───────────────────────────────────────────────────

def _str(val):
    return str(val).strip() if val is not None else ""

def _to_float(val):
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0
